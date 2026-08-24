import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import math
import statistics

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)
ws = wb.create_sheet("Excel Formulas TEXT")

#  SAMPLE DATA (Rows 1-10) 
# Column A values
a_vals = [10, 15, 7, 20, 25, 12, 8, 18, 22, 30]
for i, val in enumerate(a_vals, 1):
    ws[f'A{i}'] = val

# Column B values
b_vals = [5, 8, 12, 3, 6, 4, 9, 7, 11, 14]
for i, val in enumerate(b_vals, 1):
    ws[f'B{i}'] = val

# Column C values (scores)
c_vals = [85, 72, 65, 90, 78, 55, 88, 70, 92, 45]
for i, val in enumerate(c_vals, 1):
    ws[f'C{i}'] = val

# Column D values (names)
d_vals = ['John', 'Mary', 'Bob', 'Alice', 'Tom', 'Sue', 'Pete', 'Jane', 'Mike', 'Lisa']
for i, val in enumerate(d_vals, 1):
    ws[f'D{i}'] = val

#  CHEAT SHEET (Starting Row 13) 
current_row = 13

# Helper functions
def add_section(title):
    global current_row
    ws.cell(row=current_row, column=1, value=title)
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    current_row += 1

def add_row(func, purpose, formula, result):
    global current_row
    ws.cell(row=current_row, column=1, value=func)
    ws.cell(row=current_row, column=2, value=purpose)
    ws.cell(row=current_row, column=3, value=formula)
    ws.cell(row=current_row, column=4, value=result)
    current_row += 1

# Headers
headers = ['Function / Operator', 'Purpose', 'Formula', 'Results']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
current_row += 1

#  ARITHMETIC OPERATORS 
add_section('Arithmetic Operators')
add_row('+', 'Addition', "'=10+5", 15)
add_row('-', 'Subtraction', "'=10-5", 5)
add_row('*', 'Multiplication', "'=10*5", 50)
add_row('/', 'Division', "'=10/5", 2)
add_row('^', 'Power', "'=10^2", 100)
add_row('%', 'Percentage', "'=10%", 0.1)
add_row('&', 'Join text A1=10 and B1=5', "'=A1&\" \"&B1", "10 5")

#  COMPARISON OPERATORS 
add_section('Comparison Operators   A1 = 10           B1 = 5')
add_row('=', 'Equal to', "'=A1=B1", False)
add_row('<>', 'Not equal', "'=A1<>B1", True)
add_row('>', 'Greater than', "'=A1>B1", True)
add_row('<', 'Less than', "'=A1<B1", False)
add_row('>=', 'Greater/equal', "'=A1>=B1", True)
add_row('<=', 'Less/equal', "'=A1<=B1", False)

#  MATH & STATISTICS 
add_section('Math & Statistics')

# Calculate based on sample data
a = a_vals
b = b_vals
c = c_vals

add_row('SUM', 'Adds numbers', "'=SUM(A1:A10)", sum(a))
add_row('PRODUCT', 'Multiplies values', "'=PRODUCT(A1:A5)", a[0]*a[1]*a[2]*a[3]*a[4])
add_row('SUMPRODUCT', 'Multiplies then adds', "'=SUMPRODUCT(A1:A10,B1:B10)", sum(a[i]*b[i] for i in range(10)))
add_row('AVERAGE', 'Average', "'=AVERAGE(A1:A10)", round(sum(a)/len(a), 2))
# MEDIAN - fix for even number of values
sorted_a = sorted(a)
if len(sorted_a) % 2 == 0:
    median_val = (sorted_a[len(sorted_a)//2 - 1] + sorted_a[len(sorted_a)//2]) / 2
else:
    median_val = sorted_a[len(sorted_a)//2]
add_row('MEDIAN', 'Middle value', "'=MEDIAN(A1:A10)", median_val)
add_row('MODE.SNGL', 'Most frequent', "'=MODE.SNGL(A1:A10)", max(set(a), key=a.count))
add_row('MIN', 'Smallest', "'=MIN(A1:A10)", min(a))
add_row('MAX', 'Largest', "'=MAX(A1:A10)", max(a))
add_row('COUNT', 'Counts numbers', "'=COUNT(A1:A10)", len(a))
add_row('COUNTA', 'Counts non-empty', "'=COUNTA(A1:A10)", len(a))
add_row('COUNTBLANK', 'Counts blanks', "'=COUNTBLANK(A1:A10)", 0)
add_row('COUNTIF', 'Conditional count', "'=COUNTIF(A1:A10,\">10\")", len([x for x in a if x > 10]))
add_row('COUNTIFS', 'Multiple conditions', "'=COUNTIFS(A1:A10,\">10\",B1:B10,\">5\")", 
        len([i for i in range(10) if a[i] > 10 and b[i] > 5]))
add_row('SUMIF', 'Conditional sum', "'=SUMIF(A1:A10,\">10\",B1:B10)", 
        sum(b[i] for i in range(10) if a[i] > 10))
add_row('SUMIFS', 'Multiple-condition sum', "'=SUMIFS(C1:C10,A1:A10,\">10\",B1:B10,\">5\")",
        sum(c[i] for i in range(10) if a[i] > 10 and b[i] > 5))
add_row('AVERAGEIF', 'Conditional average', "'=AVERAGEIF(A1:A10,\">10\")",
        round(sum(x for x in a if x > 10) / len([x for x in a if x > 10]), 2))
add_row('AVERAGEIFS', 'Multiple-condition avg', "'=AVERAGEIFS(C1:C10,A1:A10,\">10\",B1:B10,\">5\")",
        round(sum(c[i] for i in range(10) if a[i] > 10 and b[i] > 5) / 
              len([i for i in range(10) if a[i] > 10 and b[i] > 5]), 2))
add_row('ROUND', 'Rounds normally', "'=ROUND(A1,2)", 10.00)
add_row('ROUNDUP', 'Rounds up', "'=ROUNDUP(A1,2)", 10.00)
add_row('ROUNDDOWN', 'Rounds down', "'=ROUNDDOWN(A1,2)", 10.00)
add_row('MROUND', 'Rounds to multiple', "'=MROUND(A1,10)", 10)
add_row('CEILING.MATH', 'Rounds up to multiple', "'=CEILING.MATH(A1,10)", 10)
add_row('FLOOR.MATH', 'Rounds down to multiple', "'=FLOOR.MATH(A1,10)", 10)
add_row('INT', 'Rounds down to integer', "'=INT(A1)", 10)
add_row('TRUNC', 'Removes decimals', "'=TRUNC(A1,2)", 10.00)
add_row('ABS', 'Absolute value', "'=ABS(A1)", 10)
add_row('MOD', 'Remainder', "'=MOD(A1,B1)", 10 % 5)
add_row('QUOTIENT', 'Whole-number division', "'=QUOTIENT(A1,B1)", 10 // 5)
add_row('POWER', 'Raises to power', "'=POWER(A1,2)", 100)
add_row('SQRT', 'Square root', "'=SQRT(A1)", round(10**0.5, 2))
add_row('RAND', 'Random decimal', "'=RAND()", '0 to 1')
add_row('RANDBETWEEN', 'Random integer', "'=RANDBETWEEN(1,100)", '1 to 100')

#  LOGICAL 
add_section('Logical')
add_row('IF', 'Tests condition', "'=IF(C1>=75,\"Pass\",\"Fail\")", 'Pass')
add_row('IFS', 'Multiple conditions', "'=IFS(C1>=90,\"A\",C1>=80,\"B\",C1>=75,\"C\")", 'B')
add_row('AND', 'All TRUE', "'=AND(A1>5,B1>5)", False)
add_row('OR', 'At least one TRUE', "'=OR(A1>5,B1>5)", True)
add_row('NOT', 'Reverses result', "'=NOT(A1>5)", False)
add_row('XOR', 'Exactly one TRUE', "'=XOR(A1>5,B1>5)", True)
add_row('IFERROR', 'Handles errors', "'=IFERROR(A1/B1,0)", 2)
add_row('IFNA', 'Handles #N/A', "'=IFNA(A1,\"Not Found\")", 10)
add_row('TRUE', 'Returns TRUE', "'=TRUE()", True)
add_row('FALSE', 'Returns FALSE', "'=FALSE()", False)
add_row('SWITCH', 'Matches values', "'=SWITCH(A1,10,\"Ten\",20,\"Twenty\",\"Other\")", 'Ten')

#  LOOKUP & REFERENCE 
add_section('Lookup & Reference')
add_row('XLOOKUP', 'Modern lookup', "'=XLOOKUP(\"John\",D1:D10,A1:A10)", 10)
add_row('VLOOKUP', 'Vertical lookup', "'=VLOOKUP(\"John\",A1:D10,4,FALSE)", 'John')
add_row('HLOOKUP', 'Horizontal lookup', "'=HLOOKUP(10,A1:D1,1,FALSE)", 10)
add_row('LOOKUP', 'Lookup value', "'=LOOKUP(\"John\",D1:D10,A1:A10)", 10)
add_row('INDEX', 'Value by position', "'=INDEX(A1:A10,3)", 7)
add_row('MATCH', 'Position of match', "'=MATCH(85,C1:C10,0)", 1)
add_row('XMATCH', 'Modern MATCH', "'=XMATCH(85,C1:C10)", 1)
add_row('CHOOSE', 'Selects by index', "'=CHOOSE(1,\"One\",\"Two\",\"Three\")", 'One')
add_row('OFFSET', 'Shifted reference', "'=OFFSET(A1,2,1)", 12)
add_row('INDIRECT', 'Reference from text', "'=INDIRECT(\"A\"&B1)", 10)
add_row('ADDRESS', 'Cell address', "'=ADDRESS(1,1)", '$A$1')
add_row('ROW', 'Row number', "'=ROW(A5)", 5)
add_row('COLUMN', 'Column number', "'=COLUMN(C1)", 3)
add_row('ROWS', 'Counts rows', "'=ROWS(A1:A10)", 10)
add_row('COLUMNS', 'Counts columns', "'=COLUMNS(A1:C1)", 3)

#  TEXT FUNCTIONS 
add_section('Text')
add_row('CONCAT', 'Joins text', "'=CONCAT(A1,B1)", '105')
add_row('CONCATENATE', 'Joins text', "'=CONCATENATE(A1,B1)", '105')
add_row('TEXTJOIN', 'Joins with delimiter', "'=TEXTJOIN(\", \",TRUE,A1:A5)", '10, 15, 7, 20, 25')
add_row('LEFT', 'Left characters', "'=LEFT(D1,3)", 'Joh')
add_row('RIGHT', 'Right characters', "'=RIGHT(D1,3)", 'ohn')
add_row('MID', 'Middle characters', "'=MID(D1,2,4)", 'ohn')
add_row('LEN', 'Character count', "'=LEN(D1)", 4)
add_row('LOWER', 'Lowercase', "'=LOWER(D1)", 'john')
add_row('UPPER', 'Uppercase', "'=UPPER(D1)", 'JOHN')
add_row('PROPER', 'Proper case', "'=PROPER(\"john doe\")", 'John Doe')
add_row('TRIM', 'Removes extra spaces', "'=TRIM(\"  John  \")", 'John')
add_row('CLEAN', 'Removes non-printing chars', "'=CLEAN(D1)", 'John')
add_row('FIND', 'Finds text, case-sensitive', "'=FIND(\"h\",D1)", 2)
add_row('SEARCH', 'Finds text, not case-sensitive', "'=SEARCH(\"H\",D1)", 2)
add_row('REPLACE', 'Replace by position', "'=REPLACE(D1,2,2,\"ABC\")", 'JABCn')
add_row('SUBSTITUTE', 'Replace matching text', "'=SUBSTITUTE(D1,\"o\",\"a\")", 'Jahn')
add_row('TEXT', 'Formats as text', "'=TEXT(A1,\"0.00\")", '10.00')
add_row('VALUE', 'Text to number', "'=VALUE(\"123\")", 123)
add_row('EXACT', 'Exact text comparison', "'=EXACT(D1,\"John\")", True)
add_row('CHAR', 'Character from code', "'=CHAR(65)", 'A')
add_row('CODE', 'Character code', "'=CODE(D1)", 74)
add_row('UNICHAR', 'Unicode character', "'=UNICHAR(9733)", '★')
add_row('UNICODE', 'Unicode code', "'=UNICODE(\"★\")", 9733)
add_row('REPT', 'Repeats text', "'=REPT(\"*\",5)", '*****')

#  DATE & TIME 
add_section('Date & Time')
today = date.today()
add_row('TODAY', 'Current date', "'=TODAY()", today.strftime('%Y-%m-%d'))
add_row('NOW', 'Current date/time', "'=NOW()", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
add_row('DATE', 'Creates date', "'=DATE(2026,8,24)", '2026-08-24')
add_row('TIME', 'Creates time', "'=TIME(14,30,0)", '14:30:00')
add_row('DAY', 'Day number', "'=DAY(DATE(2026,8,24))", 24)
add_row('MONTH', 'Month number', "'=MONTH(DATE(2026,8,24))", 8)
add_row('YEAR', 'Year', "'=YEAR(DATE(2026,8,24))", 2026)
add_row('HOUR', 'Hour', "'=HOUR(TIME(14,30,0))", 14)
add_row('MINUTE', 'Minute', "'=MINUTE(TIME(14,30,0))", 30)
add_row('SECOND', 'Second', "'=SECOND(TIME(14,30,0))", 0)
add_row('WEEKDAY', 'Weekday number', "'=WEEKDAY(DATE(2026,8,24))", 2)
add_row('WEEKNUM', 'Week number', "'=WEEKNUM(DATE(2026,8,24))", 34)
add_row('ISOWEEKNUM', 'ISO week', "'=ISOWEEKNUM(DATE(2026,8,24))", 34)
add_row('EDATE', 'Adds months', "'=EDATE(DATE(2026,8,24),3)", '2026-11-24')
add_row('EOMONTH', 'End of month', "'=EOMONTH(DATE(2026,8,24),0)", '2026-08-31')
add_row('DATEDIF', 'Date difference', "'=DATEDIF(DATE(2020,1,1),DATE(2026,8,24),\"Y\")", 6)
add_row('DAYS', 'Days between dates', "'=DAYS(DATE(2026,8,24),DATE(2020,1,1))", 2427)
add_row('NETWORKDAYS', 'Workdays', "'=NETWORKDAYS(DATE(2026,8,1),DATE(2026,8,24))", 17)
add_row('WORKDAY', 'Adds workdays', "'=WORKDAY(DATE(2026,8,24),10)", '2026-09-07')
add_row('YEARFRAC', 'Fraction of year', "'=YEARFRAC(DATE(2020,1,1),DATE(2026,8,24))", round(6.65, 2))

#  DYNAMIC ARRAYS 
add_section('Dynamic Arrays')
add_row('FILTER', 'Filters range', "'=FILTER(D1:D10,C1:C10>=75)", 'John, Mary, Alice, Tom, Pete, Mike')
add_row('SORT', 'Sorts range', "'=SORT(A1:A10,1,1)", '7,8,10,12,15,18,20,22,25,30')
add_row('SORTBY', 'Sorts by another range', "'=SORTBY(A1:A10,C1:C10,1)", 'Sorted by C values')
add_row('UNIQUE', 'Unique values', "'=UNIQUE(D1:D10)", 'John, Mary, Bob, Alice, Tom, Sue, Pete, Jane, Mike, Lisa')
add_row('SEQUENCE', 'Generates sequence', "'=SEQUENCE(10)", '1,2,3,4,5,6,7,8,9,10')
add_row('TRANSPOSE', 'Rows to columns', "'=TRANSPOSE(A1:C3)", 'Transposed')
add_row('TAKE', 'Takes rows/columns', "'=TAKE(A1:A10,5)", '10,15,7,20,25')
add_row('DROP', 'Drops rows/columns', "'=DROP(A1:A10,2)", '7,20,25,12,8,18,22,30')
add_row('CHOOSECOLS', 'Selects columns', "'=CHOOSECOLS(A1:E10,1,3)", 'Columns 1 and 3')
add_row('CHOOSEROWS', 'Selects rows', "'=CHOOSEROWS(A1:E10,1,3)", 'Rows 1 and 3')
add_row('TOCOL', 'Array to column', "'=TOCOL(A1:C3)", 'Single column')
add_row('TOROW', 'Array to row', "'=TOROW(A1:C3)", 'Single row')
add_row('VSTACK', 'Stacks vertically', "'=VSTACK(A1:B3,D1:E3)", 'Combined rows')
add_row('HSTACK', 'Stacks horizontally', "'=HSTACK(A1:B3,D1:E3)", 'Combined columns')

#  INFORMATION 
add_section('Information')
add_row('ISBLANK', 'Checks blank', "'=ISBLANK(A1)", False)
add_row('ISNUMBER', 'Checks number', "'=ISNUMBER(A1)", True)
add_row('ISTEXT', 'Checks text', "'=ISTEXT(D1)", True)
add_row('ISERROR', 'Checks any error', "'=ISERROR(A1)", False)
add_row('ISERR', 'Errors except #N/A', "'=ISERR(A1)", False)
add_row('ISNA', 'Checks #N/A', "'=ISNA(A1)", False)
add_row('ISEVEN', 'Checks even', "'=ISEVEN(A1)", True)
add_row('ISODD', 'Checks odd', "'=ISODD(A1)", False)
add_row('CELL', 'Cell information', "'=CELL(\"address\",A1)", '$A$1')
add_row('TYPE', 'Data type', "'=TYPE(A1)", 1)

#  FINANCIAL 
add_section('Financial')
add_row('PMT', 'Loan payment', "'=PMT(0.05,10,1000)", round(1000*0.05/(1-(1+0.05)**-10), 2))
add_row('PV', 'Present value', "'=PV(0.05,10,100)", round(100*(1-(1+0.05)**-10)/0.05, 2))
add_row('FV', 'Future value', "'=FV(0.05,10,100)", round(100*((1+0.05)**10-1)/0.05, 2))
add_row('NPV', 'Net present value', "'=NPV(0.05,B2:B10)", round(sum(b[1:]) * 0.95, 2))
add_row('IRR', 'Internal rate of return', "'=IRR(B2:B10)", 'IRR value')
add_row('RATE', 'Interest rate', "'=RATE(10,100,1000)", 'Rate value')
add_row('NPER', 'Number of periods', "'=NPER(0.05,100,1000)", 'Periods value')
add_row('IPMT', 'Interest payment', "'=IPMT(0.05,1,10,1000)", round(1000*0.05, 2))
add_row('PPMT', 'Principal payment', "'=PPMT(0.05,1,10,1000)", round(1000*0.05/((1+0.05)**10-1), 2))

#  COMMON ERRORS 
add_section('Common Errors')
add_row('#DIV/0!', 'Division by zero', "'=10/0", '#DIV/0!')
add_row('#N/A', 'Value unavailable', "'=XLOOKUP(\"X\",A:A,B:B)", '#N/A')
add_row('#VALUE!', 'Wrong type/value', "'=\"A\"+1", '#VALUE!')
add_row('#REF!', 'Invalid reference', "'=#REF!", '#REF!')
add_row('#NAME?', 'Unknown name/function', "'=UNKNOWN(A1)", '#NAME?')
add_row('#NUM!', 'Invalid number', "'=SQRT(-1)", '#NUM!')
add_row('#NULL!', 'Invalid intersection', "'=SUM(A1:A2 B1:B2)", '#NULL!')
add_row('#SPILL!', 'Spill blocked', "'=SEQUENCE(5)", '#SPILL!')
add_row('#CALC!', 'Calculation error', "'=FILTER(A1:A3,A1:A3>10)", '#CALC!')

#  CELL REFERENCES 
add_section('Cell References')
add_row('A1', 'Relative', "'=A1", 'Changes when copied')
add_row('$A$1', 'Absolute', "'=$A$1", 'Fixed row/column')
add_row('$A1', 'Fixed column', "'=$A1", 'Column fixed')
add_row('A$1', 'Fixed row', "'=A$1", 'Row fixed')
add_row('A1:A10', 'Range', "'=SUM(A1:A10)", sum(a))
add_row('A:A', 'Entire column', "'=SUM(A:A)", 'Column A total')
add_row('1:1', 'Entire row', "'=SUM(1:1)", 'Row 1 total')
add_row('Sheet2!A1', 'Other sheet', "'=Sheet2!A1", 'Other sheet reference')

#  USEFUL PATTERNS 
add_section('Useful Patterns')
add_row('Percentage', 'Find 10%', "'=A1*10%", 1)
add_row('Increase', 'Increase 10%', "'=A1*(1+10%)", 11)
add_row('Decrease', 'Decrease 10%', "'=A1*(1-10%)", 9)
add_row('Percentage change', 'Change %', "'=(B1-A1)/A1", -0.5)
add_row('Pass/Fail', 'Grade condition', "'=IF(C1>=75,\"Pass\",\"Fail\")", 'Pass')
add_row('Running total', 'Cumulative sum', "'=SUM($A$1:A1)", 'Running total')
add_row('Difference', 'Difference', "'=B1-A1", -5)
add_row('Age', 'Age in years', "'=DATEDIF(DATE(2000,1,1),TODAY(),\"Y\")", date.today().year - 2000)
add_row('Error handling', 'Replace errors', "'=IFERROR(A1/B1,0)", 2)

#  FORMATTING 
# Auto-adjust column widths
for col in range(1, 5):
    max_length = 0
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 50)

# Add borders to data cells (rows 1-10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for row in range(1, 11):
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border

# Freeze panes at the cheat sheet header
ws.freeze_panes = 'A14'

# Save
wb.save('Excel_Cheat_Sheet.xlsx')
print("Excel cheat sheet created successfully!")
print("File: Excel_Cheat_Sheet.xlsx")
